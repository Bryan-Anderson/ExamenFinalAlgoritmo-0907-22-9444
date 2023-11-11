import openpyxl
import tkinter  as tk

def crear_vehiculo(codigo, marca, modelo, precio, kilometraje):
    print(f" Creando producto {codigo} - {marca}")

    try:
        wb = openpyxl.load_workbook('vehiculos.xlsx')
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Código', 'Marca', 'Modelo', 'Kilometraje'])

    nuevo_vehiculos = [codigo, marca, modelo, precio, kilometraje]
    ws.append(nuevo_vehiculos)

    try:
        wb.save('vehiculos.xlsx')
        print(" Archivo Excel guardado correctamente")
    except Exception as e:
        print(f"Error al guardar el archivo Excel: {e}")


def editar_vehiculo(codigo, marca, modelo, precio, kilometraje):
    try:
        wb = openpyxl.load_workbook('vehiculos.xlsx')
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            if row[0].value == codigo:
                row[1].value = marca
                row[2].value = modelo
                row[3].value = precio
                row[4].value = kilometraje
                break
                

        wb.save('vehiculos.xlsx')
    except Exception as e:
        print(f"Error al editar el vehiculos: {e}")


def eliminar_vehiculo(codigo):
     try: 
        wb = openpyxl.load_workbook('vehiculos.xlsx')
        ws = wb.active
        filas_a_eliminar = []
        for i, row in enumerate(ws.values):
            if row[0] == codigo:
                filas_a_eliminar.append(i+1)
                for index in sorted(filas_a_eliminar, reverse=True):
                    del ws.cell(index, column=1)
                    wb.save('vehiculos.xlsx')
     except Exception as e:
         print(f"Error al eliminar el vehiculo: {e}")


def listar_vehiculo():
    try:
        wb = openpyxl.load_workbook('vehiculos.xlsx')
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2):
            data.append([valor.value for valor in row])
        return data
    
    except Exception as e:
        print(f"Error al leer los vehículos: {e}")


def guardar_vehiculo():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Codigo'
        ws['B1'] = 'Marca'
        ws['C1'] = 'Modelo'
        ws['D1'] = 'Precio'
        ws['E1'] = 'Kilometros'
        wb.save("vehiculos.xlsx")
    except Exception as e:
     print(f"Error al crear la hoja de excel: {e}")


# creame ahora la intefaz  adaptando funcione
class VentanaVehiculos:
    def __init__(self):
        self.ventana = tk.Tk()
        self.ventana.title("Manejo de Vehiculos")
        self.ventana.geometry("450x300")
        # botones
        self.btnGuardar = tk.Button(text="Guardar", command=guardar_vehiculo)
        self.btnListar = tk.Button(text="Listar", command=listar_vehiculo)
        self.btnNuevo = tk.Button(text="crear", command=crear_vehiculo)
        self.btnActualizar = tk.Button(text="Actualizar", command=eliminar_vehiculo)
        self.btnEliminar = tk.Button(text="Eliminar", command=eliminar_vehiculo)
        # etiquetas y entradas
        self.lblCodigo = tk.Label(text='Codigo')
        self.txtCodigo = tk.Entry()
        self.lblMarca = tk.Label(text='Marca')
        self.txtMarca = tk.Entry()
        self.lblModelo = tk.Label(text='Modelo')
        self.txtModelo = tk.Entry()
        self.lblPrecio = tk.Label(text='Precio')
        self.txtPrecio = tk.Entry()
        self.lblKilometros = tk.Label(text='Kilometros')
        self.txtKilometros = tk.Entry()
        
        # posicionamiento
        self.lblCodigo.grid(column=0, row=0)
        self.txtCodigo.grid(column=1, row=0)
        self.lblMarca.grid(column=0, row=1)
        self.txtMarca.grid(column=1, row=1)
        self.lblModelo.grid(column=0, row=2)
        self.txtModelo.grid(column=1, row=2)
        self.lblPrecio.grid(column=0, row=3)
        self.txtPrecio.grid(column=1, row=3)
        self.lblKilometros.grid(column=0, row=4)
        self.txtKilometros.grid(column=1, row=4)
        self.btnGuardar.grid(column=0, row=6)
        self.btnListar.grid(column=1, row=6)
        self.btnNuevo.grid(column=2, row=6)
        self.btnActualizar.grid(column=3, row=6)
        self.btnEliminar.grid(column=4, row=6)
    def main():
            ventana = VentanaVehiculos()
            ventana.ventana.mainloop()
if __name__ == "__main__":
    
    VentanaVehiculos().main()



